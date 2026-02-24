import os
import json
import tempfile
from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import LoginManager, login_required, current_user
from datetime import datetime, timedelta
import pdfplumber
import google.generativeai as genai
from dotenv import load_dotenv
import threading
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font

from models import db, User, ResearchProject
from auth import auth as auth_blueprint
from src.processor import process_financials
from src.validator import audit_financials


# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = 'super_secret_key_for_flash_messages'

# Configure Database
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///research_portal.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

with app.app_context():
    db.create_all()

# Configure Flash-Login
login_manager = LoginManager()
login_manager.login_view = 'auth.login'
login_manager.login_message_category = 'error'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Register Auth Blueprint
app.register_blueprint(auth_blueprint)

# Initialize Gemini client
# Ensure GEMINI_API_KEY is set in your environment or .env file
genai.configure(api_key=os.getenv('GEMINI_API_KEY'))

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file preserving layout and adding page markers."""
    text_content = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for idx, page in enumerate(pdf.pages):
                # Add explicit page marker for Gemini
                text_content += f"\n--- PAGE {idx + 1} ---\n"
                
                # Extract text with layout preservation
                page_text = page.extract_text(layout=True)
                if page_text:
                    text_content += page_text + "\n"
    except Exception as e:
        print(f"Error extracting text with pdfplumber: {e}")
        return None
    return text_content

@app.route('/', methods=['GET'])
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html') # Will create this new modern template

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    if current_user.auto_archive_projects:
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        old_projects = ResearchProject.query.filter(
            ResearchProject.user_id == current_user.id,
            ResearchProject.is_archived == False,
            ResearchProject.uploaded_at < thirty_days_ago
        ).all()
        for p in old_projects:
            p.is_archived = True
        if old_projects:
            db.session.commit()

    projects = ResearchProject.query.filter_by(user_id=current_user.id, is_archived=False).order_by(ResearchProject.uploaded_at.desc()).all()
    return render_template('index.html', projects=projects)

def background_process_financials(app, project_id, file_path):
    """Background task to extract and validate financials."""
    with app.app_context():
        project = ResearchProject.query.get(project_id)
        if not project:
            if os.path.exists(file_path): os.remove(file_path)
            return

        try:
            # 1. Extract Text
            extracted_text = extract_text_from_pdf(file_path)
            if extracted_text is None:
                raise Exception("Error extracting text from the PDF.")

            # 2. Call Gemini
            raw_financial_data = process_financials(extracted_text)
            
            # 3. Audit Math
            financial_data = audit_financials(raw_financial_data)
            
            # 4. Save results
            project.company_name = "Financial Report" # Simple heuristic
            project.extracted_data = json.dumps(financial_data)
            project.status = "Completed"
            
        except Exception as e:
            project.status = "Failed"
            project.extracted_data = json.dumps([{"metric": "Error", "value": str(e), "page": "-", "snippet": "-"}])
        finally:
            db.session.commit()
            if os.path.exists(file_path):
                os.remove(file_path)

@app.route('/api/upload', methods=['POST'])
@login_required
def api_upload_file():
    if 'pdf_file' not in request.files:
        return jsonify({"error": "No file part"}), 400
        
    file = request.files['pdf_file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    if file and file.filename.lower().endswith('.pdf'):
        # 1. Save file to disk temporarily
        fd, temp_path = tempfile.mkstemp(suffix='.pdf')
        with os.fdopen(fd, 'wb') as tmp:
            file.save(tmp)
            
        # 2. Create the Database record immediately as Pending
        new_project = ResearchProject(
            user_id=current_user.id,
            filename=file.filename,
            status="Pending"
        )
        db.session.add(new_project)
        db.session.commit()
        
        # 3. Spawn background thread
        thread = threading.Thread(
            target=background_process_financials,
            args=(app, new_project.id, temp_path)
        )
        thread.start()
        
        # 4. Return immediately to the frontend
        return jsonify({
            "success": True, 
            "project_id": new_project.id, 
            "status": "Pending"
        })
    else:
        return jsonify({"error": "Please upload a valid PDF file."}), 400

@app.route('/api/status/<int:project_id>', methods=['GET'])
@login_required
def check_status(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({"error": "Unauthorized"}), 403
        
    response = {
        "status": project.status,
        "project_id": project.id
    }
    
    if project.status == "Completed":
        response["redirectTo"] = url_for('view_result', project_id=project.id)
    
    return jsonify(response)

@app.route('/api/archive/<int:project_id>', methods=['POST'])
@login_required
def archive_project(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({"error": "Unauthorized"}), 403
        
    data = request.get_json() or {}
    # If explicitly passed, use it, otherwise just toggle it
    if 'is_archived' in data:
        project.is_archived = data['is_archived']
    else:
        project.is_archived = not project.is_archived
        
    db.session.commit()
    return jsonify({"success": True, "is_archived": project.is_archived})

@app.route('/api/delete/<int:project_id>', methods=['DELETE'])
@login_required
def delete_project(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({"error": "Unauthorized"}), 403
        
    db.session.delete(project)
    db.session.commit()
    return jsonify({"success": True})

@app.route('/result/<int:project_id>', methods=['GET'])
@login_required
def view_result(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Unauthorized access.')
        return redirect(url_for('dashboard'))
        
    data = []
    if project.extracted_data:
        try:
            data = json.loads(project.extracted_data)
        except json.JSONDecodeError:
            pass
            
    return render_template('result.html', data=data, filing=project)

@app.route('/api/export/<int:project_id>', methods=['GET'])
@login_required
def export_excel(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Unauthorized access.')
        return redirect(url_for('dashboard'))
        
    if not project.extracted_data:
        flash('No data to export.')
        return redirect(url_for('view_result', project_id=project_id))
        
    try:
        data = json.loads(project.extracted_data)
    except json.JSONDecodeError:
        flash('Data corruption error.')
        return redirect(url_for('view_result', project_id=project_id))
        
    if not isinstance(data, list):
        data = [data] # fallback
        
    # Build cleanly formatted list of dicts for pandas
    rows = []
    for item in data:
        if isinstance(item, dict):
            metric = item.get('metric', '-')
            v24 = item.get('value_2024', '-')
            v25 = item.get('value_2025', '-')
            growth = item.get('percentage_change', '-')
            status = item.get('status', 'Extracted')
            rows.append({
                'Metric': metric,
                '2024': v24,
                '2025': v25,
                'Growth %': growth,
                'Status': status
            })
            
    df = pd.DataFrame(rows)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Financials')
        worksheet = writer.sheets['Financials']
        
        # 1. Bold Headers and Freeze Panes
        header_font = Font(bold=True)
        for col in range(1, 6):
            worksheet.cell(row=1, column=col).font = header_font
        worksheet.freeze_panes = 'A2'
        
        # 2. Format the Growth % column based on value
        pos_fill = PatternFill(start_color="e6FFe6", end_color="e6FFe6", fill_type="solid")
        neg_fill = PatternFill(start_color="FFe6e6", end_color="FFe6e6", fill_type="solid")
        
        for row in range(2, len(rows) + 2):
            cell = worksheet.cell(row=row, column=4)
            val = cell.value
            if isinstance(val, (int, float)):
                if val > 0:
                    cell.fill = pos_fill
                elif val < 0:
                    cell.fill = neg_fill
                    
        # Widen columns
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15

        # 3. Create Summary Metadata Sheet
        summary_sheet = writer.book.create_sheet('Summary')
        summary_sheet.append(['Metadata', 'Value'])
        summary_sheet.append(['Company Name', project.company_name])
        summary_sheet.append(['Date of Extraction', datetime.utcnow().strftime('%Y-%m-%d')])
        summary_sheet.append(['Extraction Engine', 'Gemini 2.5 Flash API'])

        for col in range(1, 3):
            summary_sheet.cell(row=1, column=col).font = header_font
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 40
                    
    buffer.seek(0)
    file_name = f"{project.company_name.replace(' ', '_')}_Analysis.xlsx"
    return send_file(buffer, download_name=file_name, as_attachment=True)

@app.route('/projects', methods=['GET'])
@login_required
def projects():
    if current_user.auto_archive_projects:
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        old_projects = ResearchProject.query.filter(
            ResearchProject.user_id == current_user.id,
            ResearchProject.is_archived == False,
            ResearchProject.uploaded_at < thirty_days_ago
        ).all()
        for p in old_projects:
            p.is_archived = True
        if old_projects:
            db.session.commit()

    # Only active projects
    projects = ResearchProject.query.filter_by(user_id=current_user.id, is_archived=False).order_by(ResearchProject.uploaded_at.desc()).all()
    return render_template('projects.html', projects=projects, title="My Projects")

@app.route('/archive', methods=['GET'])
@login_required
def archive():
    # Only archived projects
    projects = ResearchProject.query.filter_by(user_id=current_user.id, is_archived=True).order_by(ResearchProject.uploaded_at.desc()).all()
    return render_template('archive.html', projects=projects, title="Archive")

@app.route('/delete_project/<int:project_id>', methods=['POST'])
@login_required
def delete_project(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Unauthorized access.', 'error')
        return redirect(url_for('dashboard'))
    
    db.session.delete(project)
    db.session.commit()
    flash('Project deleted successfully.', 'success')
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/reprocess_project/<int:project_id>', methods=['POST'])
@login_required
def reprocess_project(project_id):
    project = ResearchProject.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Unauthorized access.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        new_data = process_financials(project.extracted_text)
        project.data = new_data
        project.status = 'Completed'
        db.session.commit()
        flash('Project reprocessed successfully.', 'success')
        return redirect(url_for('view_result', project_id=project.id))
    except Exception as e:
        project.status = 'Failed'
        db.session.commit()
        flash(f'Failed to reprocess: {e}', 'error')
        return redirect(request.referrer or url_for('dashboard'))

@app.route('/settings', methods=['GET'])
@login_required
def settings():
    return render_template('settings.html', title="Settings")

@app.route('/api/settings/update', methods=['POST'])
@login_required
def update_settings():
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "Invalid payload"}), 400
        
    if 'high_contrast_mode' in data:
        current_user.high_contrast_mode = bool(data['high_contrast_mode'])
    if 'auto_archive_projects' in data:
        current_user.auto_archive_projects = bool(data['auto_archive_projects'])
    if 'email_notifications' in data:
        current_user.email_notifications = bool(data['email_notifications'])
        
    db.session.commit()
    return jsonify({"success": True})

if __name__ == '__main__':
    app.run(debug=True, port=5000)